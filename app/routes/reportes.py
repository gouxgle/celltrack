from flask import Blueprint, render_template, redirect, url_for, request, flash, make_response
from flask_login import login_required
from app.models import db, Reporte, Chip, RespxChip, Responsable, Prestadora
from sqlalchemy import text
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

bp = Blueprint('reportes', __name__, url_prefix='/reportes')


@bp.route('/')
@login_required
def index():
    # Reporte: chips con datos de facturación
    rows = db.session.execute(text("""
        SELECT
            r.nrolinea,
            r.Plan,
            r.PlanDescripcion,
            r.Importe,
            r.Status,
            r.ActivaLinea,
            ch.idchip,
            p.prestadora,
            resp.responsable,
            l.localidad,
            sec.sector
        FROM reporte r
        LEFT JOIN chip ch ON ch.nrolinea = r.nrolinea
        LEFT JOIN prestadora p ON p.idprestadora = ch.idprestadora
        LEFT JOIN respxchip rx ON rx.idchip = ch.idchip AND rx.hasta IS NULL
        LEFT JOIN responsable resp ON resp.idresponsable = rx.idresponsable
        LEFT JOIN localidad l ON l.idlocalidad = resp.idlocalidad
        LEFT JOIN sector sec ON sec.idsector = resp.idsector
        ORDER BY r.Importe DESC
    """)).fetchall()

    total = sum(r.Importe or 0 for r in rows)

    return render_template('reportes/index.html', rows=rows, total=total)


@bp.route('/por-localidad')
@login_required
def por_localidad():
    rows = db.session.execute(text("""
        SELECT
            COALESCE(l.localidad, '— Sin localidad —') AS localidad,
            COALESCE(d.distrito, '')                    AS distrito,
            ch.idchip,
            ch.nrolinea,
            ch.plan,
            ch.baja,
            p.prestadora,
            s.servicio,
            COALESCE(resp.responsable, '')              AS responsable,
            COALESCE(sec.sector, '')                    AS sector
        FROM chip ch
        LEFT JOIN prestadora p   ON p.idprestadora  = ch.idprestadora
        LEFT JOIN servicio s     ON s.idservicio     = ch.idservicio
        LEFT JOIN respxchip rx   ON rx.idchip        = ch.idchip AND rx.hasta IS NULL
        LEFT JOIN responsable resp ON resp.idresponsable = rx.idresponsable
        LEFT JOIN localidad l    ON l.idlocalidad    = resp.idlocalidad
        LEFT JOIN sector sec     ON sec.idsector     = resp.idsector
        LEFT JOIN distrito d     ON d.iddistrito     = l.iddistrito
        ORDER BY localidad, resp.responsable, ch.nrolinea
    """)).fetchall()

    # Agrupar por localidad
    from collections import OrderedDict
    grupos = OrderedDict()
    for r in rows:
        key = r.localidad
        if key not in grupos:
            grupos[key] = {'distrito': r.distrito, 'lineas': []}
        grupos[key]['lineas'].append(r)

    total_lineas  = len(rows)
    total_activas = sum(1 for r in rows if not r.baja or str(r.baja) in ('', '0000-00-00'))

    return render_template('reportes/por_localidad.html',
        grupos=grupos, total_lineas=total_lineas, total_activas=total_activas)


def _estilo_encabezado(ws, fila, columnas):
    """Aplica estilo de encabezado a una fila."""
    fill   = PatternFill('solid', fgColor='2C3E50')
    fuente = Font(color='FFFFFF', bold=True)
    borde  = Border(
        bottom=Side(style='medium', color='FFFFFF'),
    )
    for col, texto in enumerate(columnas, 1):
        c = ws.cell(row=fila, column=col, value=texto)
        c.fill   = fill
        c.font   = fuente
        c.border = borde
        c.alignment = Alignment(horizontal='center', vertical='center')


def _autoajustar(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _excel_response(wb, nombre):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return resp


@bp.route('/exportar-facturacion')
@login_required
def exportar_facturacion():
    rows = db.session.execute(text("""
        SELECT r.nrolinea, r.Plan, r.PlanDescripcion, r.Importe, r.Status,
               r.ActivaLinea, p.prestadora, resp.responsable, l.localidad, sec.sector
        FROM reporte r
        LEFT JOIN chip ch         ON ch.nrolinea       = r.nrolinea
        LEFT JOIN prestadora p    ON p.idprestadora    = ch.idprestadora
        LEFT JOIN respxchip rx    ON rx.idchip         = ch.idchip AND rx.hasta IS NULL
        LEFT JOIN responsable resp ON resp.idresponsable = rx.idresponsable
        LEFT JOIN localidad l     ON l.idlocalidad     = resp.idlocalidad
        LEFT JOIN sector sec      ON sec.idsector      = resp.idsector
        ORDER BY l.localidad, resp.responsable
    """)).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Facturación'
    ws.freeze_panes = 'A2'

    encabezados = ['Nro. Línea', 'Plan', 'Descripción Plan', 'Importe ($)',
                   'Estado', 'Alta Línea', 'Prestadora', 'Responsable', 'Localidad', 'Sector']
    _estilo_encabezado(ws, 1, encabezados)

    alt_fill = PatternFill('solid', fgColor='EBF5FB')
    for i, r in enumerate(rows, 2):
        ws.cell(i, 1, r.nrolinea)
        ws.cell(i, 2, r.Plan or '')
        ws.cell(i, 3, r.PlanDescripcion or '')
        ws.cell(i, 4, r.Importe or 0)
        ws.cell(i, 5, r.Status or '')
        ws.cell(i, 6, str(r.ActivaLinea)[:10] if r.ActivaLinea else '')
        ws.cell(i, 7, r.prestadora or '')
        ws.cell(i, 8, r.responsable or '')
        ws.cell(i, 9, r.localidad or '')
        ws.cell(i, 10, r.sector or '')
        if i % 2 == 0:
            for col in range(1, 11):
                ws.cell(i, col).fill = alt_fill

    ws.cell(len(rows) + 2, 3, 'TOTAL').font = Font(bold=True)
    ws.cell(len(rows) + 2, 4, sum(r.Importe or 0 for r in rows)).font = Font(bold=True)

    _autoajustar(ws)
    from datetime import date
    return _excel_response(wb, f'facturacion_{date.today()}.xlsx')


@bp.route('/exportar-por-localidad')
@login_required
def exportar_por_localidad():
    rows = db.session.execute(text("""
        SELECT
            COALESCE(l.localidad, '— Sin localidad —') AS localidad,
            COALESCE(d.distrito, '')                    AS distrito,
            ch.nrolinea, ch.plan, ch.baja,
            p.prestadora, s.servicio,
            COALESCE(resp.responsable, '') AS responsable,
            COALESCE(sec.sector, '')       AS sector
        FROM chip ch
        LEFT JOIN prestadora p    ON p.idprestadora    = ch.idprestadora
        LEFT JOIN servicio s      ON s.idservicio      = ch.idservicio
        LEFT JOIN respxchip rx    ON rx.idchip         = ch.idchip AND rx.hasta IS NULL
        LEFT JOIN responsable resp ON resp.idresponsable = rx.idresponsable
        LEFT JOIN localidad l     ON l.idlocalidad     = resp.idlocalidad
        LEFT JOIN sector sec      ON sec.idsector      = resp.idsector
        LEFT JOIN distrito d      ON d.iddistrito      = l.iddistrito
        ORDER BY localidad, resp.responsable, ch.nrolinea
    """)).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Líneas por Localidad'
    ws.freeze_panes = 'A2'

    encabezados = ['Localidad', 'Distrito', 'Nro. Línea', 'Prestadora',
                   'Servicio', 'Plan', 'Responsable', 'Sector', 'Estado']
    _estilo_encabezado(ws, 1, encabezados)

    alt_fill    = PatternFill('solid', fgColor='EBF5FB')
    baja_fill   = PatternFill('solid', fgColor='FADBD8')
    loc_fill    = PatternFill('solid', fgColor='D5E8D4')
    loc_fuente  = Font(bold=True, color='1A5276')

    fila          = 2
    loc_anterior  = None
    for r in rows:
        # Fila de grupo si cambia localidad
        if r.localidad != loc_anterior:
            ws.cell(fila, 1, f'{r.localidad}  —  {r.distrito}' if r.distrito else r.localidad)
            ws.cell(fila, 1).fill  = loc_fill
            ws.cell(fila, 1).font  = loc_fuente
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=9)
            fila += 1
            loc_anterior = r.localidad

        es_activa = not r.baja or str(r.baja) in ('', '0000-00-00')
        relleno   = alt_fill if fila % 2 == 0 else PatternFill()
        if not es_activa:
            relleno = baja_fill

        datos = [r.localidad, r.distrito, r.nrolinea, r.prestadora or '',
                 r.servicio or '', r.plan or '', r.responsable, r.sector,
                 'Activa' if es_activa else 'Baja']
        for col, val in enumerate(datos, 1):
            c = ws.cell(fila, col, val)
            c.fill = relleno

        fila += 1

    _autoajustar(ws)
    from datetime import date
    return _excel_response(wb, f'lineas_por_localidad_{date.today()}.xlsx')


@bp.route('/importar', methods=['GET', 'POST'])
@login_required
def importar():
    if request.method == 'POST':
        archivo = request.files.get('archivo')
        if not archivo or not archivo.filename.endswith('.xlsx'):
            flash('Seleccioná un archivo .xlsx válido.', 'danger')
            return redirect(url_for('reportes.importar'))

        import openpyxl
        try:
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active

            db.session.execute(text("DELETE FROM reporte"))

            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col     = {h: i for i, h in enumerate(headers) if h}

            insertados = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                nrolinea = str(row[col.get('nrolinea', 0)] or '').strip()
                if not nrolinea:
                    continue
                db.session.execute(text("""
                    INSERT INTO reporte (nrolinea, Bill, Plan, PlanDescripcion, Importe,
                                        PromoPlan, DescripPlan, Desde, Hasta, Status, ActivaLinea, Sim)
                    VALUES (:nrolinea, :bill, :plan, :plandesc, :importe,
                            :promo, :descrip, :desde, :hasta, :status, :activa, :sim)
                """), {
                    'nrolinea': nrolinea,
                    'bill':     row[col.get('Bill', 1)],
                    'plan':     row[col.get('Plan', 2)],
                    'plandesc': row[col.get('PlanDescripcion', 3)],
                    'importe':  row[col.get('Importe', 4)],
                    'promo':    row[col.get('PromoPlan', 5)],
                    'descrip':  row[col.get('DescripPlan', 6)],
                    'desde':    row[col.get('Desde', 7)],
                    'hasta':    row[col.get('Hasta', 8)],
                    'status':   row[col.get('Status', 9)],
                    'activa':   row[col.get('ActivaLinea', 10)],
                    'sim':      row[col.get('Sim', 11)],
                })
                insertados += 1

            db.session.commit()
            flash(f'Importación exitosa: {insertados} líneas cargadas.', 'success')
            return redirect(url_for('reportes.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al importar: {e}', 'danger')

    return render_template('reportes/importar.html')
